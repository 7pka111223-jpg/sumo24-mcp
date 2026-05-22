"""
academic_bundle.py — one-shot academic export for a SUMO24 .sumo project.

Replaces the previous workflow that required the agent to:
    1. extract data per-variable via the DTT
    2. compute aggregates by hand (TSS, removal efficiencies)
    3. fire 4 separate export_academic_results_table_docx calls
    4. write a matplotlib figure script
    5. write an openpyxl Excel script
    6. write a CSV summary

Public API:

    build_bundle(sumo_path, output_dir, *, plant_name=None,
                 scenario_label=None, include_figures=True,
                 include_excel=True, include_csv=True,
                 table_set="default") -> dict

Reads everything via sumo_offline.extract_sumo_data (no DTT required) and
writes:

    - Academic_Table1_Influent.docx
    - Academic_Table2_Effluent.docx
    - Academic_Table3_Bioreactor.docx
    - Academic_Table4_Removal.docx
    - Figure1_Influent_Characteristics.{png,pdf}
    - Figure2_Effluent_Quality.{png,pdf}
    - Figure3_Bioreactor_Profile.{png,pdf}
    - Figure4_Influent_Distributions.{png,pdf}
    - SUMO24_Simulation_Results_Academic.xlsx (6 sheets)
    - Simulation_Data_Summary.csv
"""

from __future__ import annotations

import csv
import os
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import sumo_offline as _so


# ─── Egyptian Law 48 limits used in compliance columns ─────────────────────
LAW48 = {"BOD5": 60, "TSS": 60, "TCOD": 80}


def _safe(v):
    return "" if v is None else v


def _round(v, n=3):
    if v is None or not isinstance(v, (int, float)):
        return v
    return round(v, n)


# ─── Derived effluent aggregates ───────────────────────────────────────────

def _derive_effluent(eff: Dict[str, Optional[float]]) -> Dict[str, Optional[float]]:
    """Add derived rows: SCOD = SU + SB, BOD5 ~= 0.67 * SB (rough), TCOD
    ~= SCOD + XOHO + XAOB + XNOB + XE + XU + XB + XB_e particulates × COD/VSS."""
    out = dict(eff)
    su = eff.get("SU") or 0
    sb = eff.get("SB") or 0
    out["SCOD"] = su + sb
    # Particulate COD ≈ 1.42 × volatile particulates (g COD per g VSS)
    xvss = eff.get("XVSS")
    if xvss is not None:
        out["TCOD"] = out["SCOD"] + 1.42 * xvss
    # NO2-N is included with NO3 for TKN/TN computations
    snhx = eff.get("SNHx") or 0
    sno3 = eff.get("SNO3") or 0
    sno2 = eff.get("SNO2") or 0
    out["TKN"] = snhx  # soluble + (particulate-N is small in effluent)
    out["TN"]  = snhx + sno3 + sno2
    return out


# ─── Table writers ──────────────────────────────────────────────────────────

def _influent_rows(bundle: Dict[str, Any]) -> Tuple[List[str], List[List[Any]]]:
    """Influent characteristics — design + dynamic statistics."""
    params = bundle.get("parameters") or {}
    inf = bundle.get("influent_design") or {}
    dyn = bundle.get("influent_dynamic") or {}
    # Locate the primary influent table dict
    primary_tsv = next(iter(dyn.values()), {})
    stats = primary_tsv.get("stats", {})

    def stat(col_pat: str, key: str):
        for col, s in stats.items():
            if col_pat.lower() in col.lower():
                return s.get(key)
        return None

    rows: List[List[Any]] = []
    q   = inf.get("Influent1.Q")
    bod = inf.get("Influent1.TBOD_5")
    frt = inf.get("Influent1.frTCOD_TBOD") or 0
    frv = inf.get("Influent1.frVSS_TSS")  or 0
    frs = inf.get("Influent1.frSCOD_TCOD") or 0
    tcod = bod * frt if (bod and frt) else None

    rows.append(["Flow rate (Q)",            "m³/d",   _round(q),    _round(stat("Q", "mean")), _round(stat("Q", "min")), _round(stat("Q", "max"))])
    rows.append(["BOD₅ — total",             "mg/L",   _round(bod),  _round(stat("BOD", "mean")), _round(stat("BOD", "min")), _round(stat("BOD", "max"))])
    rows.append(["Total COD (TCOD)",         "mg/L",   _round(tcod), "", "", ""])
    rows.append(["Soluble COD (SCOD)",       "mg/L",   _round(tcod * frs if (tcod and frs) else None), "", "", ""])
    rows.append(["Total Suspended Solids",   "mg/L",   _round(stat("TSS", "mean")), _round(stat("TSS", "mean")), _round(stat("TSS", "min")), _round(stat("TSS", "max"))])
    rows.append(["Volatile Suspended Solids","mg/L",   _round((stat("TSS", "mean") or 0) * frv if frv else None), "", "", ""])
    rows.append(["frTCOD/TBOD",              "—",      _round(frt), "", "", ""])
    rows.append(["frVSS/TSS",                "—",      _round(frv), "", "", ""])
    rows.append(["frSCOD/TCOD",              "—",      _round(frs), "", "", ""])

    headers = ["Parameter", "Unit", "Design", "Mean (Dyn)", "Min (Dyn)", "Max (Dyn)"]
    return headers, rows


def _effluent_rows(bundle: Dict[str, Any]) -> Tuple[List[str], List[List[Any]]]:
    """Effluent quality with Law 48 compliance."""
    eff_states = bundle.get("effluent") or {}
    ss_clarifs  = eff_states.get("steady_state") or {}
    dyn_clarifs = eff_states.get("dynamic_final") or {}
    # Use the first clarifier (typically "Clarifier3")
    cname = next(iter(ss_clarifs.keys()), None) or next(iter(dyn_clarifs.keys()), None)
    if cname is None:
        return ["Parameter", "Unit", "Effluent (SS)", "Effluent (Dyn)", "Law 48", "Compliance"], []
    eff_ss  = _derive_effluent(ss_clarifs.get(cname,  {}))
    eff_dyn = _derive_effluent(dyn_clarifs.get(cname, {}))

    def comply(ss_v, dyn_v, lim):
        if lim is None:
            return ""
        def state(v):
            if v is None:    return "—"
            return "PASS" if v <= lim else "FAIL"
        return f"{state(ss_v)} / {state(dyn_v)}"

    rows = []
    table = [
        ("BOD₅ (≈ SB)",          "mg/L",      "SB",    LAW48["BOD5"]),
        ("TCOD (computed)",      "mg/L",      "TCOD",  LAW48["TCOD"]),
        ("SCOD (computed)",      "mg/L",      "SCOD",  None),
        ("TSS (XTSS sum)",       "mg/L",      "XTSS",  LAW48["TSS"]),
        ("VSS (XVSS sum)",       "mg/L",      "XVSS",  None),
        ("NH₄-N",                "mg N/L",    "SNHx",  None),
        ("NO₃-N",                "mg N/L",    "SNO3",  None),
        ("NO₂-N",                "mg N/L",    "SNO2",  None),
        ("Total N (TN)",         "mg N/L",    "TN",    None),
        ("PO₄-P",                "mg P/L",    "SPO4",  None),
        ("DO",                   "mg/L",      "SO2",   None),
    ]
    for label, unit, key, lim in table:
        ss_v  = _round(eff_ss.get(key))
        dyn_v = _round(eff_dyn.get(key))
        rows.append([label, unit, ss_v, dyn_v, lim or "", comply(ss_v, dyn_v, lim)])

    headers = ["Parameter", "Unit", "Effluent (SS)", "Effluent (Dyn)", "Law 48 Limit", "Compliance"]
    return headers, rows


def _bioreactor_rows(bundle: Dict[str, Any]) -> Tuple[List[str], List[List[Any]]]:
    bio = bundle.get("bioreactor") or {}
    rows = []
    for unit, d in bio.items():
        rows.append([
            unit,
            _round(d.get("L_Vtrain")),
            _round(d.get("MLSS_ss"), 1),
            _round(d.get("MLSS_dyn"), 1),
            _round(d.get("DO_ss"), 3),
            _round(d.get("DO_dyn"), 3),
        ])
    headers = ["Zone", "V_train (m³)", "MLSS — SS (mg/L)", "MLSS — Dyn (mg/L)",
               "DO — SS (mg/L)", "DO — Dyn (mg/L)"]
    return headers, rows


def _removal_rows(bundle: Dict[str, Any]) -> Tuple[List[str], List[List[Any]]]:
    """Removal efficiencies SS-only (influent design → SS effluent)."""
    eff_states = bundle.get("effluent") or {}
    ss_clarifs  = eff_states.get("steady_state") or {}
    cname = next(iter(ss_clarifs.keys()), None)
    if cname is None:
        return ["Parameter", "Unit", "Influent", "Effluent (SS)", "Removal (%)"], []
    eff_ss = _derive_effluent(ss_clarifs.get(cname, {}))
    inf = bundle.get("influent_design") or {}
    params = bundle.get("parameters") or {}

    q   = inf.get("Influent1.Q")
    bod = inf.get("Influent1.TBOD_5") or 0
    frt = inf.get("Influent1.frTCOD_TBOD") or 0
    frv = inf.get("Influent1.frVSS_TSS") or 0
    frs = inf.get("Influent1.frSCOD_TCOD") or 0
    tcod = bod * frt
    scod = tcod * frs

    # Influent TSS from dynamic table mean
    dyn = bundle.get("influent_dynamic") or {}
    primary_tsv = next(iter(dyn.values()), {})
    tss_inf = next((s["mean"] for col, s in primary_tsv.get("stats", {}).items() if "TSS" in col), None)
    vss_inf = tss_inf * frv if (tss_inf and frv) else None

    def rem(inf_v, eff_v):
        if inf_v is None or eff_v is None or inf_v == 0:
            return None
        return round(100 * (inf_v - eff_v) / inf_v, 1)

    rows = [
        ["BOD₅",  "mg/L",  _round(bod),  _round(eff_ss.get("SB")),   rem(bod,  eff_ss.get("SB"))],
        ["TCOD",  "mg/L",  _round(tcod), _round(eff_ss.get("TCOD")), rem(tcod, eff_ss.get("TCOD"))],
        ["SCOD",  "mg/L",  _round(scod), _round(eff_ss.get("SCOD")), rem(scod, eff_ss.get("SCOD"))],
        ["TSS",   "mg/L",  _round(tss_inf), _round(eff_ss.get("XTSS")), rem(tss_inf, eff_ss.get("XTSS"))],
        ["VSS",   "mg/L",  _round(vss_inf), _round(eff_ss.get("XVSS")), rem(vss_inf, eff_ss.get("XVSS"))],
    ]
    headers = ["Parameter", "Unit", "Influent", "Effluent (SS)", "Removal (%)"]
    return headers, rows


# ─── Figure generators ─────────────────────────────────────────────────────

def _figures(bundle: Dict[str, Any], output_dir: Path,
             scenario_label: str) -> List[str]:
    """Generate the 4 publication figures (PNG + PDF each). Returns list
    of written paths. Skips silently if matplotlib not available."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np
    except Exception:
        return []

    plt.rcParams.update({
        "font.family": "serif", "font.size": 10,
        "axes.labelsize": 11, "axes.titlesize": 11,
        "figure.dpi": 300,
        "axes.spines.top": False, "axes.spines.right": False,
    })

    written: List[str] = []
    # Pull dynamic influent series
    dyn = bundle.get("influent_dynamic") or {}
    primary_name, primary = next(iter(dyn.items()), (None, {}))
    rows = []
    if primary_name and primary:
        sumo_path = Path(bundle["sumo_path"])
        with zipfile.ZipFile(sumo_path, "r") as zf:
            if primary_name in zf.namelist():
                text = zf.read(primary_name).decode("utf-8", "replace")
                _, rows = _so.read_influent_tsv(text)
    headers = primary.get("headers", [])
    days  = np.array([(r.get(headers[0]) or 0) / 86400000 for r in rows]) if (rows and headers) else np.array([])

    def col(name_pat: str):
        for h in headers[1:]:
            if name_pat.lower() in h.lower():
                vals = [r.get(h) for r in rows]
                return h, vals
        return None, []
    q_hdr, q_vals = col("Q")
    bod_hdr, bod_vals = col("BOD")
    tss_hdr, tss_vals = col("TSS")

    # ─── FIGURE 1: Influent characterisation ─────────────────────────
    if len(days) and q_vals:
        fig, axes = plt.subplots(3, 1, figsize=(7.5, 7.5), sharex=True)
        fig.subplots_adjust(hspace=0.08)
        ax = axes[0]
        ax.step(days, np.array(q_vals, dtype=float) / 1000, where="post", color="#2980b9")
        ax.set_ylabel("Q (10³ m³/d)")
        ax.text(0.01, 0.93, "(a) Flow rate", transform=ax.transAxes, fontweight="bold")
        if bod_vals:
            ax = axes[1]
            pts = [(d, v) for d, v in zip(days, bod_vals) if v is not None]
            if pts:
                ax.scatter([p[0] for p in pts], [p[1] for p in pts],
                           s=20, color="#c0392b")
            ax.axhline(60, ls=":", color="grey", label="Law 48 (60 mg/L)")
            ax.set_ylabel("BOD₅ (mg/L)")
            ax.text(0.01, 0.93, "(b) Influent BOD₅", transform=ax.transAxes, fontweight="bold")
            ax.legend(loc="upper right", frameon=False)
        if tss_vals:
            ax = axes[2]
            pts = [(d, v) for d, v in zip(days, tss_vals) if v is not None]
            if pts:
                ax.scatter([p[0] for p in pts], [p[1] for p in pts],
                           s=12, color="#27ae60", marker="s")
            ax.axhline(60, ls=":", color="grey", label="Law 48 (60 mg/L)")
            ax.set_ylabel("TSS (mg/L)")
            ax.set_xlabel("Simulation time (days)")
            ax.text(0.01, 0.93, "(c) Influent TSS", transform=ax.transAxes, fontweight="bold")
            ax.legend(loc="upper right", frameon=False)
        fig.suptitle(f"Influent Characteristics — {scenario_label}", fontsize=10.5, y=0.98)
        for ext in ("png", "pdf"):
            p = output_dir / f"Figure1_Influent_Characteristics.{ext}"
            fig.savefig(str(p), dpi=300, bbox_inches="tight")
            written.append(str(p))
        plt.close(fig)

    # ─── FIGURE 2: Effluent quality bar chart ────────────────────────
    eff_states = bundle.get("effluent") or {}
    ss_clarifs  = eff_states.get("steady_state") or {}
    dyn_clarifs = eff_states.get("dynamic_final") or {}
    cname = next(iter(ss_clarifs.keys()), None)
    if cname:
        eff_ss  = _derive_effluent(ss_clarifs.get(cname,  {}))
        eff_dyn = _derive_effluent(dyn_clarifs.get(cname, {}))
        fig, axes = plt.subplots(1, 3, figsize=(9, 4.5))
        fig.subplots_adjust(wspace=0.38)
        # Panel a: organic matter
        keys_a = ["SB", "TCOD", "SCOD"]
        labels_a = ["BOD₅", "TCOD", "SCOD"]
        ax = axes[0]
        x = np.arange(len(keys_a)); w = 0.35
        ax.bar(x - w/2, [eff_ss.get(k)  or 0 for k in keys_a], w, color="#2980b9", label="SS")
        ax.bar(x + w/2, [eff_dyn.get(k) or 0 for k in keys_a], w, color="#27ae60", label="Dyn")
        ax.set_xticks(x); ax.set_xticklabels(labels_a)
        ax.set_ylabel("Concentration (mg/L)")
        ax.set_title("(a) Organic Matter", fontweight="bold")
        ax.legend(frameon=False, fontsize=8)
        # Panel b: nitrogen
        keys_b = ["SNHx", "SNO3", "SNO2", "TN"]
        labels_b = ["NH₄-N", "NO₃-N", "NO₂-N", "TN"]
        ax = axes[1]
        x = np.arange(len(keys_b))
        ax.bar(x - w/2, [eff_ss.get(k)  or 0 for k in keys_b], w, color="#2980b9", label="SS")
        ax.bar(x + w/2, [eff_dyn.get(k) or 0 for k in keys_b], w, color="#27ae60", label="Dyn")
        ax.set_xticks(x); ax.set_xticklabels(labels_b)
        ax.set_ylabel("mg N/L")
        ax.set_title("(b) Nitrogen", fontweight="bold")
        ax.legend(frameon=False, fontsize=8)
        # Panel c: solids & P
        keys_c = ["XTSS", "XVSS", "SPO4"]
        labels_c = ["TSS", "VSS", "PO₄-P"]
        ax = axes[2]
        x = np.arange(len(keys_c))
        ax.bar(x - w/2, [eff_ss.get(k)  or 0 for k in keys_c], w, color="#2980b9", label="SS")
        ax.bar(x + w/2, [eff_dyn.get(k) or 0 for k in keys_c], w, color="#27ae60", label="Dyn")
        ax.set_xticks(x); ax.set_xticklabels(labels_c)
        ax.set_ylabel("mg/L")
        ax.set_title("(c) Solids & P", fontweight="bold")
        ax.legend(frameon=False, fontsize=8)
        fig.suptitle(f"Simulated Effluent Quality — {scenario_label}", fontsize=10.5, y=0.99)
        for ext in ("png", "pdf"):
            p = output_dir / f"Figure2_Effluent_Quality.{ext}"
            fig.savefig(str(p), dpi=300, bbox_inches="tight")
            written.append(str(p))
        plt.close(fig)

    # ─── FIGURE 3: Bioreactor MLSS + DO ──────────────────────────────
    bio = bundle.get("bioreactor") or {}
    if bio:
        names = list(bio.keys())
        mlss_ss  = [bio[n].get("MLSS_ss")  or 0 for n in names]
        mlss_dyn = [bio[n].get("MLSS_dyn") or 0 for n in names]
        do_ss    = [bio[n].get("DO_ss")    or 0 for n in names]
        do_dyn   = [bio[n].get("DO_dyn")   or 0 for n in names]
        fig, axes = plt.subplots(1, 2, figsize=(10, 4))
        fig.subplots_adjust(wspace=0.32)
        x = np.arange(len(names)); w = 0.4
        ax = axes[0]
        ax.bar(x - w/2, mlss_ss,  w, color="#2980b9", label="SS")
        ax.bar(x + w/2, mlss_dyn, w, color="#27ae60", label="Dyn")
        ax.set_xticks(x); ax.set_xticklabels(names, rotation=40, ha="right", fontsize=8)
        ax.set_ylabel("MLSS (mg/L)")
        ax.set_title("(a) MLSS profile", fontweight="bold")
        ax.legend(frameon=False, fontsize=8)
        ax = axes[1]
        ax.bar(x - w/2, do_ss,  w, color="#2980b9", label="SS")
        ax.bar(x + w/2, do_dyn, w, color="#27ae60", label="Dyn")
        ax.set_xticks(x); ax.set_xticklabels(names, rotation=40, ha="right", fontsize=8)
        ax.set_ylabel("DO (mg/L)")
        ax.axhline(2.0, color="#c0392b", ls="--", lw=1, label="Aerobic set point")
        ax.set_title("(b) DO profile", fontweight="bold")
        ax.legend(frameon=False, fontsize=8)
        fig.suptitle(f"Bioreactor Profile — {scenario_label}", fontsize=10.5, y=0.99)
        for ext in ("png", "pdf"):
            p = output_dir / f"Figure3_Bioreactor_Profile.{ext}"
            fig.savefig(str(p), dpi=300, bbox_inches="tight")
            written.append(str(p))
        plt.close(fig)

    # ─── FIGURE 4: Influent BOD & TSS histograms ────────────────────
    if bod_vals or tss_vals:
        fig, axes = plt.subplots(1, 2, figsize=(8, 4))
        fig.subplots_adjust(wspace=0.32)
        bvs = [v for v in (bod_vals or []) if v is not None]
        tvs = [v for v in (tss_vals or []) if v is not None]
        if bvs:
            ax = axes[0]
            ax.hist(bvs, bins=16, color="#c0392b", alpha=0.75, edgecolor="white")
            ax.axvline(sum(bvs) / len(bvs), color="#7b241c", ls="--", label=f"Mean = {sum(bvs)/len(bvs):.0f}")
            ax.set_xlabel("BOD₅ (mg/L)"); ax.set_ylabel("Frequency")
            ax.set_title("(a) Influent BOD₅", fontweight="bold")
            ax.legend(frameon=False, fontsize=8)
        if tvs:
            ax = axes[1]
            ax.hist(tvs, bins=22, color="#27ae60", alpha=0.75, edgecolor="white")
            ax.axvline(sum(tvs) / len(tvs), color="#1d8348", ls="--", label=f"Mean = {sum(tvs)/len(tvs):.0f}")
            ax.set_xlabel("TSS (mg/L)"); ax.set_ylabel("Frequency")
            ax.set_title("(b) Influent TSS", fontweight="bold")
            ax.legend(frameon=False, fontsize=8)
        fig.suptitle(f"Influent Distributions — {scenario_label}", fontsize=10.5, y=0.99)
        for ext in ("png", "pdf"):
            p = output_dir / f"Figure4_Influent_Distributions.{ext}"
            fig.savefig(str(p), dpi=300, bbox_inches="tight")
            written.append(str(p))
        plt.close(fig)

    return written


# ─── Excel workbook ────────────────────────────────────────────────────────

def _excel(bundle: Dict[str, Any], output_dir: Path,
           scenario_label: str) -> Optional[str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    BLUE_DARK = "1F497D"; BLUE_MID = "4472C4"; BLUE_LIGHT = "D6E4F0"
    GREY = "F2F2F2"; WHITE = "FFFFFF"; GREEN = "E8F8F0"; RED = "FDECEA"

    def fill(c): return PatternFill("solid", fgColor=c)
    def border():
        s = Side(style="thin"); return Border(left=s, right=s, top=s, bottom=s)
    def hdr(ws, row, cols, texts, bg=BLUE_DARK):
        for col, txt in zip(cols, texts):
            c = ws.cell(row=row, column=col, value=txt)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = fill(bg); c.border = border()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    def datarow(ws, row, vals, bg=WHITE):
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = Font(size=10); c.fill = fill(bg); c.border = border()
            if isinstance(v, (int, float)):
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = "0.000"
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
    def title(ws, row, text, ncols):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=12, color=BLUE_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill(BLUE_LIGHT)
        ws.row_dimensions[row].height = 22

    wb = Workbook(); wb.remove(wb.active)

    # Sheet 1: Influent
    h, rows = _influent_rows(bundle)
    ws = wb.create_sheet("1. Influent")
    ws.sheet_view.showGridLines = False
    title(ws, 1, f"Table 1. Influent — {scenario_label}", len(h))
    hdr(ws, 2, list(range(1, len(h)+1)), h)
    for i, r in enumerate(rows):
        datarow(ws, 3 + i, r, bg=(GREY if i % 2 == 0 else WHITE))
    for i, w in enumerate([32, 10, 14, 16, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2: Effluent
    h, rows = _effluent_rows(bundle)
    ws = wb.create_sheet("2. Effluent")
    ws.sheet_view.showGridLines = False
    title(ws, 1, f"Table 2. Effluent Quality — {scenario_label}", len(h))
    hdr(ws, 2, list(range(1, len(h)+1)), h)
    for i, r in enumerate(rows):
        bg = GREY if i % 2 == 0 else WHITE
        if "FAIL" in (r[-1] or ""):
            bg = RED
        elif "PASS" in (r[-1] or "") and "FAIL" not in (r[-1] or ""):
            bg = GREEN
        datarow(ws, 3 + i, r, bg=bg)
    for i, w in enumerate([28, 10, 16, 18, 14, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 3: Bioreactor
    h, rows = _bioreactor_rows(bundle)
    ws = wb.create_sheet("3. Bioreactor")
    ws.sheet_view.showGridLines = False
    title(ws, 1, f"Table 3. Bioreactor — {scenario_label}", len(h))
    hdr(ws, 2, list(range(1, len(h)+1)), h)
    for i, r in enumerate(rows):
        datarow(ws, 3 + i, r, bg=(GREY if i % 2 == 0 else WHITE))
    for i, w in enumerate([24, 14, 18, 18, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 4: Removal
    h, rows = _removal_rows(bundle)
    ws = wb.create_sheet("4. Removal Efficiencies")
    ws.sheet_view.showGridLines = False
    title(ws, 1, f"Table 4. Removal Efficiencies — {scenario_label}", len(h))
    hdr(ws, 2, list(range(1, len(h)+1)), h)
    for i, r in enumerate(rows):
        datarow(ws, 3 + i, r, bg=(GREY if i % 2 == 0 else WHITE))
    for i, w in enumerate([24, 10, 14, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 5: Biomass
    bm = bundle.get("biomass") or {}
    ws = wb.create_sheet("5. Biomass")
    ws.sheet_view.showGridLines = False
    title(ws, 1, f"Table 5. Bioreactor Biomass (SS, averaged across zones) — {scenario_label}", 3)
    hdr(ws, 2, [1, 2, 3], ["Symbol", "Concentration (mg/L)", "Description"])
    descriptions = {
        "XOHO":   "Ordinary heterotrophs",
        "XAOB":   "Ammonia-oxidising bacteria",
        "XNOB":   "Nitrite-oxidising bacteria",
        "XPAO":   "Polyphosphate-accumulating organisms",
        "XAMX":   "Anammox",
        "XACP":   "Acetate-producing bacteria",
        "XAMETO": "Aceticlastic methanogens",
        "XALGAE": "Algae",
        "XB":     "Slowly biodegradable particulate",
        "XB_e":   "Endogenous biodegradable particulate",
        "XE":     "Endogenous residue",
        "XU":     "Unbiodegradable particulate",
        "XINORG": "Inorganic particulate",
        "XPP":    "Polyphosphate",
        "XTSS":   "Total suspended solids (sum)",
        "XVSS":   "Volatile suspended solids (sum)",
    }
    row = 3
    for sym, val in bm.items():
        if val is None: continue
        datarow(ws, row, [sym, _round(val, 3), descriptions.get(sym, "")],
                bg=(GREY if row % 2 == 0 else WHITE))
        row += 1
    for i, w in enumerate([12, 22, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 6: Dynamic Influent
    dyn = bundle.get("influent_dynamic") or {}
    primary_name, primary = next(iter(dyn.items()), (None, {}))
    ws = wb.create_sheet("6. Dynamic Influent")
    ws.sheet_view.showGridLines = False
    title(ws, 1, f"Table 6. Dynamic Influent — {primary_name or 'n/a'}", 4)
    if primary_name:
        sumo_path = Path(bundle["sumo_path"])
        with zipfile.ZipFile(sumo_path) as zf:
            if primary_name in zf.namelist():
                text = zf.read(primary_name).decode("utf-8", "replace")
                hdrs, rows_data = _so.read_influent_tsv(text)
                hdr(ws, 2, list(range(1, len(hdrs)+1)), hdrs)
                for i, rec in enumerate(rows_data):
                    vals = [rec.get(h) for h in hdrs]
                    datarow(ws, 3 + i, vals, bg=(GREY if i % 2 == 0 else WHITE))
                for i, w in enumerate([16] * len(hdrs), 1):
                    ws.column_dimensions[get_column_letter(i)].width = w
                ws.freeze_panes = "A3"

    # Tab colours
    for w_idx, ws in enumerate(wb.worksheets):
        ws.sheet_properties.tabColor = ["1F497D", "2E75B6", "2E75B6",
                                         "70AD47", "C55A11", "A9A9A9"][w_idx]

    out = output_dir / "SUMO24_Simulation_Results_Academic.xlsx"
    wb.save(str(out))
    return str(out)


# ─── CSV summary ──────────────────────────────────────────────────────────

def _csv(bundle: Dict[str, Any], output_dir: Path,
         scenario_label: str) -> str:
    out = output_dir / "Simulation_Data_Summary.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([f"# Simulation Data Summary — {scenario_label}"])
        w.writerow([])
        for label, (h, rows) in [
            ("=== Influent ===",     _influent_rows(bundle)),
            ("=== Effluent ===",     _effluent_rows(bundle)),
            ("=== Bioreactor ===",   _bioreactor_rows(bundle)),
            ("=== Removal % ===",    _removal_rows(bundle)),
        ]:
            w.writerow([label])
            w.writerow(h)
            for r in rows:
                w.writerow(r)
            w.writerow([])
    return str(out)


# ─── Top-level entry ──────────────────────────────────────────────────────

def build_bundle(
    sumo_path: str,
    output_dir: str,
    *,
    plant_name: Optional[str] = None,
    scenario_label: Optional[str] = None,
    include_figures: bool = True,
    include_excel: bool = True,
    include_csv: bool = True,
    table_set: str = "default",
    docx_writer=None,
) -> Dict[str, Any]:
    """Generate the full academic export from a .sumo file.

    `docx_writer` is an optional callable matching the signature of
    server._exp_academic_results_docx. If supplied, the 4 booktabs Word
    tables are also generated; otherwise the CSV / Excel cover the same
    data and the docx step is silently skipped.
    """
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    bundle = _so.extract_sumo_data(sumo_path)
    if not bundle.get("ok"):
        return {"ok": False, "error": bundle.get("error"), "stage": "extract"}

    scenario_label = scenario_label or (plant_name or Path(sumo_path).stem)
    written: Dict[str, Any] = {"tables_docx": [], "figures": [], "excel": None, "csv": None}

    # Tables (docx) — optional, via injected writer
    if docx_writer is not None:
        for idx, (label, (h, rows)) in enumerate([
            ("Influent",   _influent_rows(bundle)),
            ("Effluent",   _effluent_rows(bundle)),
            ("Bioreactor", _bioreactor_rows(bundle)),
            ("Removal",    _removal_rows(bundle)),
        ], start=1):
            p = out_dir / f"Academic_Table{idx}_{label}.docx"
            try:
                docx_writer(
                    headers=h,
                    rows=[[("" if v is None else v) for v in r] for r in rows],
                    output_path=str(p),
                    caption=f"{label} summary — {scenario_label}",
                    table_number=idx,
                    column_units=None,
                    title=None,
                    footnote="Generated by export_academic_bundle.",
                )
                written["tables_docx"].append(str(p))
            except Exception:
                pass

    if include_figures:
        written["figures"] = _figures(bundle, out_dir, scenario_label)

    if include_excel:
        written["excel"] = _excel(bundle, out_dir, scenario_label)

    if include_csv:
        written["csv"] = _csv(bundle, out_dir, scenario_label)

    return {
        "ok": True,
        "sumo_path": sumo_path,
        "output_dir": str(out_dir),
        "scenario_label": scenario_label,
        "files": written,
        "summary": (
            f"Bundle: {len(written['tables_docx'])} Word tables, "
            f"{len(written['figures'])} figures, "
            f"{'1' if written['excel'] else '0'} Excel workbook, "
            f"{'1' if written['csv'] else '0'} CSV."
        ),
    }
