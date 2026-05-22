"""
pipeline/seeding.py
───────────────────
Stage 2 helper: pre-populate the Excel template's Treatment_Processes and
Connections sheets from the parsed schematic so the user only needs to fill
in numbers — unit names and connections are already locked in.

This eliminates the entire class of "WAS unit missing from Treatment_Processes"
errors that caused the Qaha1 session failures.
"""
from __future__ import annotations

import pathlib
import shutil
from datetime import datetime
from typing import Any

# ── Unit-type → process-type mapping ─────────────────────────────────────────
# Maps schematic unit_type keys to the human-readable strings used in the
# Treatment_Processes.process_type column.

_PROCESS_TYPE: dict[str, str] = {
    "influent":           "Influent",
    "effluent":           "Effluent",
    "ras_flow":           "RAS Flow",
    "was_flow":           "WAS Flow",
    "screen":             "Screen",
    "grit_chamber":       "Grit Chamber",
    "primary_clarifier":  "Primary Clarifier",
    "oxidation_ditch":    "Oxidation Ditch",
    "aeration_tank":      "Aeration Tank",
    "anoxic_zone":        "Anoxic Zone",
    "anaerobic_zone":     "Anaerobic Zone",
    "mbbr":               "MBBR",
    "mbr":                "MBR",
    "secondary_clarifier":"Secondary Clarifier",
    "sludge_splitter":    "Sludge Splitter",
    "sludge_thickener":   "Sludge Thickener",
    "anaerobic_digester": "Anaerobic Digester",
    "aerobic_digester":   "Aerobic Digester",
    "centrifuge":         "Centrifuge",
}

_CATEGORY: dict[str, str] = {
    "influent":           "boundary",
    "effluent":           "boundary",
    "ras_flow":           "boundary",
    "was_flow":           "boundary",
    "screen":             "pretreatment",
    "grit_chamber":       "pretreatment",
    "primary_clarifier":  "pretreatment",
    "oxidation_ditch":    "biological",
    "aeration_tank":      "biological",
    "anoxic_zone":        "biological",
    "anaerobic_zone":     "biological",
    "mbbr":               "biological",
    "mbr":                "biological",
    "secondary_clarifier":"separation",
    "sludge_splitter":    "physical",
    "sludge_thickener":   "sludge",
    "anaerobic_digester": "sludge",
    "aerobic_digester":   "sludge",
    "centrifuge":         "sludge",
}

_STAGE: dict[str, str] = {
    "influent":           "influent",
    "effluent":           "effluent",
    "ras_flow":           "recycle",
    "was_flow":           "waste",
    "screen":             "pretreatment",
    "grit_chamber":       "pretreatment",
    "primary_clarifier":  "pretreatment",
    "oxidation_ditch":    "biological",
    "aeration_tank":      "biological",
    "anoxic_zone":        "biological",
    "anaerobic_zone":     "biological",
    "mbbr":               "biological",
    "mbr":                "biological",
    "secondary_clarifier":"secondary",
    "sludge_splitter":    "recycle",
    "sludge_thickener":   "sludge",
    "anaerobic_digester": "sludge",
    "aerobic_digester":   "sludge",
    "centrifuge":         "sludge",
}

_KINETIC: dict[str, str] = {
    "oxidation_ditch":    "ASM2d",
    "aeration_tank":      "ASM2d",
    "anoxic_zone":        "ASM2d",
    "anaerobic_zone":     "ASM2d",
    "mbbr":               "ASM2d",
    "mbr":                "ASM2d",
}

_RECYCLE_TYPES = {"ras", "was", "recycle", "internal_recycle", "sludge_recycle"}


def _unit_name_by_id(schematic: dict, uid: str) -> str:
    for u in schematic.get("units", []):
        if u.get("id") == uid:
            return u.get("name", uid)
    return uid


def seed_excel_from_schematic(
    template_path: "str | pathlib.Path",
    schematic: dict,
    kinetic_model: str = "ASM2d",
    backup: bool = True,
) -> dict:
    """
    Write schematic units → Treatment_Processes sheet, and
    schematic streams → Connections sheet.

    Only the columns that are derivable from the schematic are written;
    user-editable columns (volumes, areas, influent concentrations, etc.)
    are left untouched if they already have values — so re-seeding after
    the user has entered data is safe.

    Returns a result dict with ok, rows_written, and any warnings.
    """
    try:
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return {"ok": False, "error": "openpyxl is not installed"}

    p = pathlib.Path(template_path)
    if not p.exists():
        return {"ok": False, "error": f"Template not found: {template_path}"}

    if backup:
        ts = datetime.utcnow().strftime("%Y%m%dT%H%M%S")
        backup_path = p.parent / f"{p.stem}.bak.{ts}{p.suffix}"
        shutil.copy2(p, backup_path)

    try:
        wb = openpyxl.load_workbook(p)
    except Exception as ex:
        return {"ok": False, "error": f"Cannot open workbook: {ex}"}

    warnings: list[str] = []
    units = schematic.get("units", [])
    streams = schematic.get("streams", [])

    # ── Treatment_Processes sheet ────────────────────────────────────────────
    if "Treatment_Processes" not in wb.sheetnames:
        warnings.append("Treatment_Processes sheet missing — creating it")
        wb.create_sheet("Treatment_Processes")
    ws_tp = wb["Treatment_Processes"]

    # Ensure header row exists
    headers_tp = [
        "order_index", "instance_name", "process_type",
        "category", "stage", "kinetic_model",
    ]
    if ws_tp.max_row < 1 or ws_tp.cell(1, 1).value is None:
        for col, h in enumerate(headers_tp, 1):
            ws_tp.cell(1, col).value = h

    # Build a lookup: instance_name → row (for existing data preservation)
    existing_names: dict[str, int] = {}
    for row in ws_tp.iter_rows(min_row=2, values_only=False):
        name_cell = row[1] if len(row) > 1 else None
        if name_cell and name_cell.value:
            existing_names[str(name_cell.value)] = name_cell.row

    tp_written = 0
    for i, u in enumerate(units, start=1):
        uname = u.get("name", f"Unit{i}")
        utype = u.get("type", "")

        if uname in existing_names:
            # Preserve existing row — only fill blank cells
            r = existing_names[uname]
        else:
            r = ws_tp.max_row + 1

        def _set_if_blank(col: int, value: Any) -> None:
            cell = ws_tp.cell(r, col)
            if cell.value is None:
                cell.value = value

        _set_if_blank(1, i)
        ws_tp.cell(r, 2).value = uname   # instance_name always written (primary key)
        _set_if_blank(3, _PROCESS_TYPE.get(utype, utype))
        _set_if_blank(4, _CATEGORY.get(utype, "physical"))
        _set_if_blank(5, _STAGE.get(utype, ""))
        _set_if_blank(6, _KINETIC.get(utype, kinetic_model) if utype in _KINETIC else "")
        tp_written += 1

    # ── Connections sheet ────────────────────────────────────────────────────
    if "Connections" not in wb.sheetnames:
        warnings.append("Connections sheet missing — creating it")
        wb.create_sheet("Connections")
    ws_conn = wb["Connections"]

    headers_conn = ["from_unit", "to_unit", "stream_label", "stream_type", "recycle_type"]
    if ws_conn.max_row < 1 or ws_conn.cell(1, 1).value is None:
        for col, h in enumerate(headers_conn, 1):
            ws_conn.cell(1, col).value = h

    # Build existing-connection lookup to avoid duplicates
    existing_conns: set[tuple] = set()
    for row in ws_conn.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            existing_conns.add((str(row[0]), str(row[1])))

    conn_written = 0
    for s in streams:
        from_name = _unit_name_by_id(schematic, s.get("from", ""))
        to_name   = _unit_name_by_id(schematic, s.get("to", ""))
        if not from_name or not to_name:
            continue
        if (from_name, to_name) in existing_conns:
            continue  # preserve existing row

        stype = s.get("type", "liquid")
        label = s.get("label") or f"{from_name}_to_{to_name}"
        recycle = stype if stype in _RECYCLE_TYPES else ""

        r = ws_conn.max_row + 1
        ws_conn.cell(r, 1).value = from_name
        ws_conn.cell(r, 2).value = to_name
        ws_conn.cell(r, 3).value = label
        ws_conn.cell(r, 4).value = stype
        ws_conn.cell(r, 5).value = recycle
        existing_conns.add((from_name, to_name))
        conn_written += 1

    # ── Data-validation dropdown on Connections A & B columns ───────────────
    if units:
        n = len(units) + 1
        dv = DataValidation(
            type="list",
            formula1=f"=Treatment_Processes!$B$2:$B${n}",
            allow_blank=True,
            showDropDown=False,
        )
        ws_conn.add_data_validation(dv)
        max_rows = max(len(streams) * 2 + 10, 50)
        dv.add(f"A2:A{max_rows}")
        dv.add(f"B2:B{max_rows}")

    try:
        wb.save(p)
    except Exception as ex:
        return {"ok": False, "error": f"Cannot save workbook: {ex}"}

    return {
        "ok": True,
        "template_path": str(p),
        "treatment_processes_written": tp_written,
        "connections_written": conn_written,
        "warnings": warnings,
    }
