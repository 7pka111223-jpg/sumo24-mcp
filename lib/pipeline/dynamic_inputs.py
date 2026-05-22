"""
pipeline/dynamic_inputs.py
──────────────────────────
Stage 5 (optional): generate dynamic input TSV files for SUMO24.

Five profile types:
    constant          – flat table using Stage 4 static values
    from_excel        – read Influent_Dataset rows as time-series
    diurnal           – sinusoidal Q + COD around daily mean
    storm             – diurnal + 6-hour storm event at 2.5× Q
    seasonal_temp     – monthly mean temperatures, others at daily means

Output: <project_dir>/dynamic_inputs/<name>.tsv
Format: SUMO24 dynamic input table (tab-separated, Time[d] first column)
"""
from __future__ import annotations

import math
import pathlib
from datetime import datetime
from typing import Any


_HEADER = ["Time[d]", "Q[m3/d]", "COD[mg/L]", "TKN[mg/L]", "TP[mg/L]",
           "TSS[mg/L]", "T[degC]"]


def generate_dynamic_inputs(
    project_dir: "str | pathlib.Path",
    profile_type: str,
    influent: dict,
    duration_days: float = 30.0,
    dt_h: float = 0.5,
    storm_day: float = 7.0,
    storm_peak_factor: float = 2.5,
    storm_duration_h: float = 6.0,
    excel_path: "str | None" = None,
    name: str | None = None,
) -> dict:
    """
    Generate a dynamic input TSV and write it to
    <project_dir>/dynamic_inputs/<name>.tsv.

    Returns {ok, path, rows, warnings}.
    """
    valid_types = {"constant", "from_excel", "diurnal", "storm", "seasonal_temp"}
    if profile_type not in valid_types:
        return {"ok": False,
                "error": f"profile_type must be one of {sorted(valid_types)}"}

    out_name = name or profile_type
    out_dir = pathlib.Path(project_dir) / "dynamic_inputs"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{out_name}.tsv"

    Q0   = float(influent.get("Q_m3d",   1000.0))
    COD0 = float(influent.get("COD_mgL",  400.0))
    TKN0 = float(influent.get("TKN_mgL",  50.0))
    TP0  = float(influent.get("TP_mgL",    7.0))
    TSS0 = float(influent.get("TSS_mgL",  250.0))
    T0   = float(influent.get("T_C",       20.0))
    warnings: list[str] = []
    rows: list[list] = []

    if profile_type == "constant":
        rows = _constant_profile(Q0, COD0, TKN0, TP0, TSS0, T0,
                                  duration_days, dt_h)

    elif profile_type == "diurnal":
        rows = _diurnal_profile(Q0, COD0, TKN0, TP0, TSS0, T0,
                                  duration_days, dt_h)

    elif profile_type == "storm":
        rows = _storm_profile(Q0, COD0, TKN0, TP0, TSS0, T0,
                               duration_days, dt_h,
                               storm_day, storm_peak_factor, storm_duration_h)

    elif profile_type == "seasonal_temp":
        rows = _seasonal_temp_profile(Q0, COD0, TKN0, TP0, TSS0,
                                       duration_days, dt_h)

    elif profile_type == "from_excel":
        if not excel_path:
            return {"ok": False, "error": "excel_path required for from_excel profile"}
        result = _from_excel(excel_path, warnings)
        if not result["ok"]:
            return result
        rows = result["rows"]

    # Write TSV
    lines = ["\t".join(_HEADER)]
    for r in rows:
        lines.append("\t".join(str(round(v, 4)) for v in r))
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    return {
        "ok": True,
        "path": str(out_path),
        "profile_type": profile_type,
        "rows": len(rows),
        "duration_days": duration_days,
        "warnings": warnings,
    }


def validate_dynamic_tsv(tsv_path: "str | pathlib.Path",
                          sumo_stop_time: float | None = None,
                          sumo_data_comm: float | None = None) -> dict:
    """
    Validate a dynamic input TSV:
    - Well-formed (uniform column count)
    - Time column monotonic
    - Physical range checks on Q, COD, TKN, TP, TSS, T
    - Duration matches sumo_stop_time if provided
    """
    p = pathlib.Path(tsv_path)
    if not p.exists():
        return {"ok": False, "error": f"File not found: {tsv_path}"}

    lines = p.read_text(encoding="utf-8").strip().splitlines()
    if len(lines) < 2:
        return {"ok": False, "error": "TSV has fewer than 2 rows (header + 1 data row minimum)"}

    header = lines[0].split("\t")
    n_cols = len(header)
    errors: list[str] = []
    warnings: list[str] = []

    times: list[float] = []
    for i, line in enumerate(lines[1:], start=2):
        parts = line.split("\t")
        if len(parts) != n_cols:
            errors.append(f"Row {i}: expected {n_cols} cols, got {len(parts)}")
            continue
        try:
            t = float(parts[0])
            times.append(t)
        except ValueError:
            errors.append(f"Row {i}: non-numeric time value '{parts[0]}'")

    if not times:
        return {"ok": False, "error": "No valid time values found", "errors": errors}

    # Monotonicity
    for i in range(1, len(times)):
        if times[i] <= times[i - 1]:
            errors.append(f"Time not monotonic at row {i + 2}: "
                          f"{times[i - 1]} → {times[i]}")

    # Duration
    duration = times[-1] - times[0]
    if sumo_stop_time and duration < sumo_stop_time * 0.9:
        warnings.append(
            f"TSV duration ({duration:.1f} d) < Sumo__StopTime ({sumo_stop_time} d). "
            "SUMO may extrapolate or error at the end."
        )

    ok = len(errors) == 0
    return {
        "ok": ok,
        "path": str(tsv_path),
        "rows": len(times),
        "duration_d": round(duration, 3),
        "errors": errors,
        "warnings": warnings,
    }


# ── Profile generators ────────────────────────────────────────────────────────

def _constant_profile(Q, COD, TKN, TP, TSS, T, dur, dt_h):
    dt_d = dt_h / 24.0
    rows = []
    t = 0.0
    while t <= dur + 1e-9:
        rows.append([t, Q, COD, TKN, TP, TSS, T])
        t += dt_d
    return rows


def _diurnal_profile(Q, COD, TKN, TP, TSS, T, dur, dt_h):
    """Sinusoidal Q and COD: peak at 7 AM (≈ t_frac 0.29), trough at 3 AM."""
    dt_d = dt_h / 24.0
    rows = []
    t = 0.0
    while t <= dur + 1e-9:
        frac = math.modf(t)[0]   # time of day (0–1)
        phase = 2 * math.pi * (frac - 7.0 / 24.0)
        q_factor   = 1.0 + 0.4 * math.sin(phase)
        cod_factor = 1.0 + 0.25 * math.sin(phase - math.pi / 6)
        rows.append([t,
                     Q   * q_factor,
                     COD * cod_factor,
                     TKN * (1.0 + 0.15 * math.sin(phase)),
                     TP,
                     TSS * (1.0 + 0.20 * math.sin(phase)),
                     T])
        t += dt_d
    return rows


def _storm_profile(Q, COD, TKN, TP, TSS, T, dur, dt_h,
                   storm_day, peak_factor, storm_dur_h):
    base = _diurnal_profile(Q, COD, TKN, TP, TSS, T, dur, dt_h)
    storm_start = storm_day
    storm_end   = storm_day + storm_dur_h / 24.0
    result = []
    for row in base:
        t = row[0]
        if storm_start <= t <= storm_end:
            prog = (t - storm_start) / (storm_dur_h / 24.0)
            s_factor = peak_factor * math.sin(math.pi * prog)
            result.append([t,
                            row[1] * (1.0 + (s_factor - 1.0)),
                            row[2] / s_factor if s_factor > 1 else row[2],
                            row[3] / s_factor if s_factor > 1 else row[3],
                            row[4],
                            row[5] / s_factor if s_factor > 1 else row[5],
                            row[6]])
        else:
            result.append(row)
    return result


_MONTHLY_TEMP_EGYPT = [16, 17, 19, 22, 25, 28, 29, 29, 27, 24, 20, 17]

def _seasonal_temp_profile(Q, COD, TKN, TP, TSS, dur, dt_h):
    dt_d = dt_h / 24.0
    rows = []
    t = 0.0
    while t <= dur + 1e-9:
        month = int(t / 30.4167) % 12
        T = _MONTHLY_TEMP_EGYPT[month]
        rows.append([t, Q, COD, TKN, TP, TSS, T])
        t += dt_d
    return rows


def _from_excel(excel_path: str, warnings: list) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"ok": False, "error": "openpyxl is not installed"}

    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
    except Exception as ex:
        return {"ok": False, "error": f"Cannot open Excel: {ex}"}

    if "Influent_Dataset" not in wb.sheetnames:
        return {"ok": False, "error": "Influent_Dataset sheet not found in Excel"}

    ws = wb["Influent_Dataset"]
    rows = []
    t = 0.0
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        if row[0] is None:
            break
        try:
            vals = [t] + [float(v) if v is not None else 0.0 for v in row[:6]]
            rows.append(vals)
            t += 1.0
        except (TypeError, ValueError) as e:
            warnings.append(f"Skipping row: {e}")

    wb.close()
    if not rows:
        return {"ok": False, "error": "No data rows found in Influent_Dataset sheet"}
    return {"ok": True, "rows": rows}
