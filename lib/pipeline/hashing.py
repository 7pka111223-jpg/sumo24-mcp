"""
pipeline/hashing.py
───────────────────
Stable, content-addressed hashes for each pipeline artefact type.

Rules:
- Timestamps and volatile fields are EXCLUDED so hashes are stable
  across round-trips.
- Each hash function is deterministic: same logical content → same hash.
- Hashes are hex-encoded SHA-256 strings (64 chars).
"""
from __future__ import annotations

import hashlib
import json
import pathlib


# ── HTML schematic ────────────────────────────────────────────────────────────

def hash_schematic(html_path: "str | pathlib.Path") -> str:
    """
    Hash only the JSON block embedded in <script id="schematic-data">.
    Strips volatile fields (created, modified) before hashing so that
    re-opening and re-saving the builder without changing topology doesn't
    mark the stage stale.
    """
    import re
    raw = pathlib.Path(html_path).read_text(encoding="utf-8", errors="replace")

    # Comment-safe extraction: blank out <!-- ... --> before matching
    _COMMENT_RE = re.compile(r"<!--.*?-->", re.DOTALL)
    blanked = _COMMENT_RE.sub(lambda m: " " * len(m.group()), raw)

    _TAG_RE = re.compile(
        r'<script\s+id="schematic-data"\s+type="application/json"\s*>(.*?)</script>',
        re.DOTALL,
    )
    m = _TAG_RE.search(blanked)
    if not m:
        # Fall back to hashing the whole file (stage will mark as stale if
        # the schematic block is ever added)
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()

    # Locate the span in the *original* text (offsets are the same because
    # the blanker preserves length)
    start, end = m.span(1)
    json_text = raw[start:end].strip()

    try:
        obj = json.loads(json_text)
    except json.JSONDecodeError:
        # Malformed JSON — hash the raw text so we still get a stable hash
        return hashlib.sha256(json_text.encode("utf-8")).hexdigest()

    # Strip volatile metadata fields
    for volatile in ("created", "modified", "lastSaved"):
        obj.get("meta", {}).pop(volatile, None)

    canonical = json.dumps(obj, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


# ── Excel template ────────────────────────────────────────────────────────────

def hash_excel(xlsx_path: "str | pathlib.Path") -> str:
    """
    Hash every non-empty cell value across all sheets (sorted sheet order).
    Uses data_only=False so formula text is hashed rather than cached values,
    which avoids spurious hash changes when formulas are recalculated.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        # openpyxl unavailable — fall back to raw bytes hash
        return hashlib.sha256(pathlib.Path(xlsx_path).read_bytes()).hexdigest()

    wb = load_workbook(xlsx_path, data_only=False, read_only=True)
    h = hashlib.sha256()
    for ws_name in sorted(wb.sheetnames):
        ws = wb[ws_name]
        h.update(ws_name.encode("utf-8"))
        for row in ws.iter_rows(values_only=True):
            row = tuple(row)
            # Drop trailing Nones (avoids phantom empty columns)
            while row and row[-1] is None:
                row = row[:-1]
            if any(c is not None for c in row):
                h.update(repr(row).encode("utf-8"))
    wb.close()
    return h.hexdigest()


# ── JSON files ────────────────────────────────────────────────────────────────

def hash_json(path: "str | pathlib.Path") -> str:
    """
    Hash a JSON file with sorted keys so dict ordering differences don't
    produce different hashes.
    """
    obj = json.loads(pathlib.Path(path).read_text(encoding="utf-8"))
    canonical = json.dumps(obj, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


# ── Build pack directory ──────────────────────────────────────────────────────

def hash_pack(pack_dir: "str | pathlib.Path") -> str:
    """
    Hash an entire directory by concatenating (relative-path, file-bytes)
    pairs in sorted order.  Returns a stable hash of the whole pack.
    """
    pd = pathlib.Path(pack_dir)
    h = hashlib.sha256()
    for f in sorted(pd.rglob("*")):
        if f.is_file():
            h.update(str(f.relative_to(pd)).encode("utf-8"))
            h.update(f.read_bytes())
    return h.hexdigest()
